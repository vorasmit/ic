import collections

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def __init__(self):
        pass

    def get_workbook(self, workbook=None):
        if not workbook:
            workbook = openpyxl.Workbook()

        return workbook

    def make_xlsx(self, *args, **kwargs):
        """
        Make xlsx file
        :param workbook - Object of excel file/ workbook
        :param sheet_name - name for the worksheet
        :param filters - A data dictionary to added in sheet
        :param merged_headers - A dict of List
            @example: {
                'label': [column1, colum2]
            }
        :param headers: A List of dictionary (cell properties will be optional)
        :param data: A list of dictionary to append data to sheet
        :param file_name: Name of excel file/workbook
        """
        # Create workbook if not passed or return existing workbook
        wb = self.get_workbook(kwargs.get("workbook"))

        if "workbook" not in kwargs:
            kwargs["workbook"] = wb

        # ToDo: Create a new sheet
        cs = Worksheet(**kwargs)
        cs.create_worksheet()

        self.save_workbook(wb, kwargs.get("file_name"))
        return wb

    def save_workbook(self, wb, file_name):
        """Save workbook"""
        if file_name[-4:] != ".xlsx":
            file_name = f"{file_name}.xlsx"

        wb.save(file_name)

    def remove_worksheet(self, wb, sheet_name):
        """Remove worksheet"""
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])


class Worksheet:
    def __init__(self, **kwargs):
        self.row_dimension = 1
        self.column_dimension = 1

        for (k, v) in kwargs.items():
            setattr(self, k, v)

    def create_worksheet(self):
        """Create worksheet"""
        self.ws = self.workbook.create_sheet(self.sheet_name)

        if hasattr(self, "filters"):
            self.add_data(self.filters)

        if hasattr(self, "merged_headers"):
            self.add_data(self.merged_headers, merge=True)

        if hasattr(self, "headers"):
            self.add_data(self.headers, is_header=True)

        self.add_data(self.data, is_data=True)

        if hasattr(self, "add_totals") and self.add_totals:
            self.add_total_row(self.data)

    def add_data(
        self, data, is_header=False, is_data=False, merge=False, is_total=False
    ):
        """Adds header data to the sheet"""
        parsed_data = self.parse_data(data, is_header)

        for i, row in enumerate(parsed_data, 1):
            for j, val in enumerate(row):
                cell = self.ws.cell(row=self.row_dimension, column=j + 1)

                if merge:
                    self.append_merged_header()
                else:
                    self.apply_style(
                        self.row_dimension,
                        j + 1,
                        column_index=j,
                        is_header=is_header,
                        is_data=is_data,
                        is_total=is_total,
                    )
                    cell.value = val
            self.row_dimension += 1

    def add_total_row(self, data):
        """Add total row to the sheet"""
        counter = collections.Counter()

        for row in data:
            counter.update(row)

        total_row = list(counter.values())
        total_row.pop(0)
        total_row.insert(0, "Totals")

        self.replace_string(total_row, skip_value="Totals")
        self.add_data(total_row, is_header=True, is_total=True)

    def append_merged_header(self):
        for key, value in self.merged_headers.items():
            start_column = self.get_column_index(value[0])
            end_column = self.get_column_index(value[1])

            range = self.get_range(
                start_row=self.row_dimension,
                start_column=start_column,
                end_row=self.row_dimension,
                end_column=end_column,
            )

            self.ws.cell(row=self.row_dimension, column=start_column).value = key

            self.ws.merge_cells(range)

            self.apply_style(
                self.row_dimension,
                start_column,
                column_index=start_column,
                is_header=True,
            )

    def get_column_index(self, column_name):
        """Get column index from column name"""
        for (idx, field) in enumerate(self.headers, 1):
            if field["fieldname"] == column_name:
                return idx

    def apply_style(
        self, row, column, column_index, is_header=False, is_data=False, is_total=False
    ):
        """Apply style to cell"""
        self.get_properties(column_index, is_header, is_data, is_total)

        cell = self.ws.cell(row=row, column=column)
        cell.font = Font(
            name=self.font_family,
            size=self.font_size,
            bold=self.bold,
        )
        cell.alignment = Alignment(
            horizontal=self.horizontal_align,
            vertical=self.vertical_align,
            wrap_text=self.wrap_text,
        )
        cell.number_format = self.format if self.format else "General"

        if self.bg_color:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.bg_color,
            )
        self.ws.column_dimensions[get_column_letter(column)].width = self.width
        self.ws.row_dimensions[row].height = self.height

    def get_properties(
        self, column=None, is_header=False, is_data=False, is_total=False
    ):
        """Get all properties defined in a header for cell"""
        if not column:
            column = self.column_dimension

        properties = self.headers[column]

        self.font_family = "Calibri"
        self.font_size = 9
        self.bold = True
        self.align_header = "center"
        self.align_data = "general"
        self.format = "General"
        self.width = 20
        self.height = 20
        self.vertical_align = "bottom"
        self.wrap_text = False
        self.bg_color = None

        for (k, v) in properties.items():
            setattr(self, k, v)

        if is_header:
            self.horizontal_align = self.align_header
            self.bg_color = self.bg_color
            self.font_size = self.font_size
            self.vertical_align = "center"
            self.wrap_text = True
            self.height = 30
            if is_total:
                self.horizontal_align = self.align_data
                self.height = 20
        elif is_data:
            self.horizontal_align = self.align_data
            self.bg_color = self.bg_color_data
            self.bold = False
        else:
            self.horizontal_align = self.align_data
            self.bg_color = None

    def parse_data(self, data, is_header=False):
        """Convert data to List of Lists"""
        csv_list = []

        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, list):
                    csv_list.append(list(data.keys()))
                    return csv_list
                else:
                    csv_list.append([key, value])

        if isinstance(data, list):
            csv_list = self.build_csv_array(data, is_header)

        return csv_list

    def build_csv_array(self, data, is_header=False):
        """Builds a csv data array"""
        csv_array = []
        csv_row = []

        for row in data:
            if isinstance(row, dict):
                if is_header:
                    csv_row.append(row.get("label"))
                else:
                    csv_array.append(list(row.values()))
            else:
                csv_array.append(data)
                break

        if len(csv_row) >= 1:
            csv_array.append(csv_row)

        return csv_array

    def get_range(self, start_row, start_column, end_row, end_column):
        start_column_letter = get_column_letter(start_column)
        end_column_letter = get_column_letter(end_column)

        range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

        return range

    def replace_string(self, data_list, skip_value=None):
        """Replace string with empty string in Total Row"""

        for i, data in enumerate(data_list):
            if isinstance(data, str):
                if data == skip_value:
                    continue
                data_list[i] = ""
        return data_list
